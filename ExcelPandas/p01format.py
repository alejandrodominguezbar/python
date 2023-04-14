from openpyxl import load_workbook
from openpyxl.styles import Font, Fill

# Load the workbook and select the active worksheet
wb = load_workbook('filep01output.xlsx')
ws = wb.active

# Set the font style of cell A1 to bold
ws['A1'].font = Font(bold=True)

a1 = ws['A1']
d4 = ws['D4']
ft = Font(color="FF0000")
a1.font = ft
d4.font = ft

# Set the background color of cells A2 to A4 to yellow
for i in range(2, 18):
   ws[f'A{i}'].font = ft

# Save the workbook to a file
wb.save('p01outputformatted.xlsx')