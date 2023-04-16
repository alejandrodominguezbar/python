# Importar bibliotecas
import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

# Crear una ventana raíz
root = tk.Tk()
root.withdraw()

# Preguntar al usuario si desea agregar información a la hoja "bacalao"
result = messagebox.askyesno("Question", "Do you want to add data to the 'bacalao' sheet?")

if result:
    # Pedir al usuario que seleccione un archivo
    file_path = filedialog.askopenfilename()

    # Leer el archivo Excel seleccionado en un DataFrame de pandas
    df = pd.read_excel(file_path)

    # Agregar nueva información al DataFrame
    df["Nuevo dato 1"] = "Información adicional"
    df["Nuevo dato 2"] = 123

    # Abrir el archivo Excel existente
    workbook = openpyxl.load_workbook(file_path)

    # Crear una nueva hoja con nombre "bacalao" si no existe
    if "bacalao" not in workbook.sheetnames:
        workbook.create_sheet("bacalao")

    # Seleccionar la hoja "bacalao"
    sheet = workbook["bacalao"]

    # Escribir los datos del DataFrame en la hoja "bacalao"
    for row in dataframe_to_rows(df, index=False, header=True):
        sheet.append(row)

    # Guardar los cambios en el archivo Excel
    workbook.save(file_path)

    # Mostrar un mensaje de confirmación
    messagebox.showinfo("Info", "Data added to the 'bacalao' sheet in the selected file.")
else:
    # Mostrar un mensaje de cancelación
    messagebox.showinfo("Info", "No data was added to the 'bacalao' sheet.")